"""Excel 匯出：xlsx 7 欄，含 skipped / NULL corrected_text 段落。

Self-contained fixture — follows project pattern (test_correction_router.py).
"""
from __future__ import annotations

import io

import pytest
from app.services.correction.excel_exporter import session_to_excel_bytes
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _seed_session(
    db: Session,
    *,
    api_key_id: int,
    segments: list[dict],
) -> int:
    db.execute(
        text(
            "INSERT INTO audio_files "
            "(api_key_id, original_name, storage_path, file_size, duration_sec) "
            "VALUES (:a, 'excel.wav', '/tmp/excel.wav', 1024, 5.0)"
        ),
        {"a": api_key_id},
    )
    audio_id = int(db.execute(
        text("SELECT id FROM audio_files WHERE api_key_id = :a ORDER BY id DESC LIMIT 1"),
        {"a": api_key_id},
    ).scalar_one())

    db.execute(
        text(
            "INSERT INTO transcriptions "
            "(api_key_id, source, model_name, model_version, transcript_text, duration_sec) "
            "VALUES (:a, 'upload', 'm', 'v1', 'orig', 5.0)"
        ),
        {"a": api_key_id},
    )
    tx_id = int(db.execute(
        text("SELECT id FROM transcriptions WHERE api_key_id = :a ORDER BY id DESC LIMIT 1"),
        {"a": api_key_id},
    ).scalar_one())
    db.execute(
        text("UPDATE audio_files SET transcription_id = :t WHERE id = :a"),
        {"t": tx_id, "a": audio_id},
    )

    db.execute(
        text(
            "INSERT INTO correction_sessions (api_key_id, transcription_id, name) "
            "VALUES (:a, :t, 'excel-test')"
        ),
        {"a": api_key_id, "t": tx_id},
    )
    sess_id = int(db.execute(
        text("SELECT id FROM correction_sessions WHERE api_key_id = :a ORDER BY id DESC LIMIT 1"),
        {"a": api_key_id},
    ).scalar_one())

    for s in segments:
        db.execute(
            text(
                "INSERT INTO correction_segments "
                "(session_id, segment_index, start_sec, end_sec, "
                "original_text, corrected_text, speaker_label, is_skipped) "
                "VALUES (:sid, :idx, :st, :en, :orig, :corr, :spk, :skip)"
            ),
            {
                "sid": sess_id,
                "idx": s["index"],
                "st": s["start"],
                "en": s["end"],
                "orig": s["original"],
                "corr": s.get("corrected"),
                "spk": s.get("speaker"),
                "skip": s.get("skipped", False),
            },
        )
    db.commit()
    return sess_id


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def excel_setup(db_session: Session):
    db_session.execute(
        text(
            "TRUNCATE api_keys, transcriptions, correction_sessions, "
            "correction_segments, audio_files CASCADE"
        )
    )
    db_session.execute(
        text(
            "INSERT INTO api_keys (key_hash, lookup_prefix, name, scopes) "
            "VALUES ('$argon2id$dummy', '1234567890abcdef', 'excel-key', '{asr:write}')"
        )
    )
    api_key_id = int(db_session.execute(
        text("SELECT id FROM api_keys WHERE name = 'excel-key'")
    ).scalar_one())
    db_session.commit()
    return db_session, api_key_id


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_session_to_excel_columns_and_rows(excel_setup) -> None:
    """7 欄表頭正確；所有段落均輸出（含 skipped / NULL corrected）；speaker NULL → 空字串。"""
    db, api_key_id = excel_setup
    sess_id = _seed_session(
        db,
        api_key_id=api_key_id,
        segments=[
            {"index": 0, "start": 0.0, "end": 1.5, "original": "原文 A",
             "corrected": "校正 A", "speaker": "S0", "skipped": False},
            {"index": 1, "start": 1.5, "end": 3.0, "original": "原文 B",
             "corrected": "校正 B", "speaker": None, "skipped": False},
            {"index": 2, "start": 3.0, "end": 4.0, "original": "原文 C",
             "corrected": None, "speaker": "S1", "skipped": True},
        ],
    )
    raw = session_to_excel_bytes(db, sess_id, api_key_id=api_key_id)
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # 第一列 = 表頭
    assert rows[0] == ("段落", "開始", "結束", "語者", "原文", "校正", "已跳過")
    # 共 3 段 + 1 表頭
    assert len(rows) == 1 + 3
    # 第 2 段 speaker_label NULL → 空字串（openpyxl values_only=True 回傳空儲存格為 None）
    assert rows[2][3] in (None, "")
    # 第 3 段 is_skipped = True
    assert rows[3][6] is True
